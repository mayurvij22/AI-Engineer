import os
import openpyxl
import pypdf

# Paths
# Paths - resolve relative to this file so it works both locally and in Vercel serverless
DATA_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "data_pack"))
EXCEL_PATH = os.path.join(DATA_DIR, "ParcelPilot_Assessment_Data.xlsx")

# In-memory database
accounts = []
orders = []
tickets = []
documents = {}  # filename -> full text

def load_data():
    global accounts, orders, tickets, documents
    
    # 1. Load Excel Sheets
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"Excel file not found at {EXCEL_PATH}")
        
    wb = openpyxl.load_workbook(EXCEL_PATH)
    
    # Accounts
    accounts_sheet = wb["accounts"]
    accounts_rows = list(accounts_sheet.iter_rows(values_only=True))
    accounts_headers = [str(h) for h in accounts_rows[0]]
    accounts = []
    for r in accounts_rows[1:]:
        if any(v is not None for v in r):
            accounts.append(dict(zip(accounts_headers, r)))
            
    # Orders
    orders_sheet = wb["orders"]
    orders_rows = list(orders_sheet.iter_rows(values_only=True))
    orders_headers = [str(h) for h in orders_rows[0]]
    orders = []
    for r in orders_rows[1:]:
        if any(v is not None for v in r):
            orders.append(dict(zip(orders_headers, r)))
            
    # Tickets
    tickets_sheet = wb["tickets"]
    tickets_rows = list(tickets_sheet.iter_rows(values_only=True))
    tickets_headers = [str(h) for h in tickets_rows[0]]
    tickets = []
    for r in tickets_rows[1:]:
        if any(v is not None for v in r):
            tickets.append(dict(zip(tickets_headers, r)))
            
    print(f"Loaded {len(accounts)} accounts, {len(orders)} orders, {len(tickets)} tickets.")

    # 2. Load PDF Documents
    pdf_files = [
        "01_Support_Policy_v3_CURRENT.pdf",
        "02_Support_Policy_v2_DEPRECATED.pdf",
        "03_Cancellation_and_Service_Credit_SOP_v4.pdf",
        "04_Product_Operations_Guide_and_Known_Issues.pdf",
        "05_Northstar_Logistics_Enterprise_Agreement.pdf",
        "06_LumenWorks_Service_Agreement.pdf"
    ]
    
    documents = {}
    for pdf_name in pdf_files:
        pdf_path = os.path.join(DATA_DIR, pdf_name)
        if os.path.exists(pdf_path):
            try:
                reader = pypdf.PdfReader(pdf_path)
                text_content = []
                for page in reader.pages:
                    extracted = page.extract_text() or ""
                    text_content.append(extracted)
                documents[pdf_name] = "\n".join(text_content)
            except Exception as e:
                print(f"Warning: Failed to load PDF {pdf_name}: {e}")
        else:
            print(f"Warning: PDF {pdf_name} not found in data_pack.")
            
    print(f"Loaded and indexed {len(documents)} PDF documents.")

# Initialize on import
load_data()

def search_documents(query: str, account_id: str = None) -> list:
    """
    Searches the loaded documents.
    If account_id is provided, filters custom agreements to only include the customer's agreement.
    Returns a list of dicts: {'filename': str, 'score': float, 'snippet': str, 'full_text': str}
    """
    query_words = [w.lower() for w in query.split() if len(w) > 2]
    if not query_words:
        query_words = [query.lower()]
        
    results = []
    
    for filename, text in documents.items():
        # Scoping rule:
        # If it's a signed agreement, check if it matches the customer's account
        is_agreement = "Agreement" in filename or "Enterprise" in filename
        if is_agreement and account_id:
            # Let's check which agreement belongs to this account
            account = next((a for a in accounts if a["account_id"] == account_id), None)
            if not account or account.get("contract_file") != filename:
                # This customer is not authorized to read this agreement!
                continue
        elif is_agreement and not account_id:
            # Internal context: authorized to read all agreements
            pass
            
        # Basic scoring: count occurrences of query words
        score = 0
        text_lower = text.lower()
        for word in query_words:
            score += text_lower.count(word)
            
        if score > 0:
            # Extract a matching snippet
            first_word_idx = -1
            for word in query_words:
                first_word_idx = text_lower.find(word)
                if first_word_idx != -1:
                    break
                    
            if first_word_idx != -1:
                start = max(0, first_word_idx - 100)
                end = min(len(text), first_word_idx + 250)
                snippet = text[start:end].replace("\n", " ")
                if start > 0:
                    snippet = "..." + snippet
                if end < len(text):
                    snippet = snippet + "..."
            else:
                snippet = text[:200].replace("\n", " ") + "..."
                
            results.append({
                "filename": filename,
                "score": score,
                "snippet": snippet.strip(),
                "full_text": text
            })
            
    # Sort by score descending
    results.sort(key=lambda x: x["score"], reverse=True)
    return results

def get_accounts():
    return accounts

def get_orders(account_id: str = None):
    if account_id:
        return [o for o in orders if o["account_id"] == account_id]
    return orders

def get_tickets(account_id: str = None):
    if account_id:
        return [t for t in tickets if t["account_id"] == account_id]
    return tickets

def add_ticket(ticket: dict):
    tickets.insert(0, ticket)
    return ticket

def update_ticket(ticket_id: str, updates: dict):
    for t in tickets:
        if t["ticket_id"] == ticket_id:
            t.update(updates)
            return t
    return None

def update_order(order_id: str, updates: dict):
    for o in orders:
        if o["order_id"] == order_id:
            o.update(updates)
            return o
    return None
